import streamlit as st
import pandas as pd
import datetime
import calendar
import io
from api import get_cert_expired, get_cases, get_case_statistics, get_equipment_maintenance, get_attendance_by_month,get_worklogs_by_case_id, get_all_cases_statistics
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

def export_salary_to_excel(attendance, month):
    """
    將薪資資料輸出為Excel檔案
    
    Args:
        attendance: 出勤資料
        month: 月份
        
    Returns:
        bytes: Excel檔案的二進位資料
    """
    # 確保必要的套件
    import pandas as pd
    import io
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # 解析月份
    year, month_num = map(int, month.split('-'))
    _, last_day = calendar.monthrange(year, month_num)
    days_in_month = list(range(1, last_day + 1))
    
    # 創建日期欄位
    columns = ["姓名/日期"] + [str(day) for day in days_in_month] + ["合計"]
    
    # 創建表格資料
    table_data = []
    
    # 將每位員工的出勤資料轉換為表格格式
    for emp in attendance['employees']:
        row = [emp['name']]  # 第一欄是員工姓名
        
        # 建立日期對應的工時字典
        hours_by_day = {}
        for record in emp['daily_records']:
            day = int(record['date'].split('-')[2])  # 取出日期中的「日」
            hours_by_day[day] = record['hours']
        
        # 填充每一天的工時
        for day in days_in_month:
            hours = hours_by_day.get(day, 0)
            # 如果工時大於0，顯示工時值，否則顯示空白
            row.append(hours if hours > 0 else "")
        
        # 最後一欄是總計
        row.append(emp['total_hours'])
        
        table_data.append(row)
    
    # 添加合計行
    total_row = ["合計"]
    for day in days_in_month:
        # 計算每一天的工時總和
        day_total = 0
        for emp in table_data:
            day_index = day  # 日期對應的索引
            if day_index < len(emp) and emp[day_index] != "":
                day_total += float(emp[day_index])
        
        total_row.append(day_total if day_total > 0 else "")
    
    # 最後一格是所有員工總工時
    grand_total = sum(float(emp[-1]) for emp in table_data if emp[-1] != "")
    total_row.append(grand_total)
    
    table_data.append(total_row)
    
    # 創建 DataFrame
    df = pd.DataFrame(table_data, columns=columns)
    
    # 創建 Excel 檔案
    output = io.BytesIO()
    
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 將資料輸出到Excel
            df.to_excel(writer, sheet_name=f'{month}月薪資單', index=False)
            
            # 取得工作表以進行格式設定
            workbook = writer.book
            worksheet = writer.sheets[f'{month}月薪資單']
            
            # 設定欄位寬度
            for i, column in enumerate(df.columns):
                column_letter = get_column_letter(i + 1)
                if i == 0:  # 姓名欄位
                    worksheet.column_dimensions[column_letter].width = 15
                elif i == len(df.columns) - 1:  # 合計欄位
                    worksheet.column_dimensions[column_letter].width = 10
                else:  # 日期欄位
                    worksheet.column_dimensions[column_letter].width = 4
            
            # 設定標題列格式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            header_alignment = Alignment(horizontal='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 設定合計行格式
            total_row_num = len(table_data) + 1  # +1 因為Excel的行從1開始，且有標題行
            for cell in worksheet[total_row_num]:
                cell.font = Font(bold=True)
            
            # 設定數值格式
            for row in worksheet.iter_rows(min_row=2, max_row=total_row_num):
                for cell in row:
                    if cell.value not in ["", None] and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'
    except Exception as e:
        # 輸出錯誤訊息以協助除錯
        st.error(f"創建 Excel 檔案時發生錯誤: {str(e)}")
        # 返回空的二進位資料
        return io.BytesIO().getvalue()
    
    # 重置指針位置並返回
    output.seek(0)
    return output.getvalue()

def display_case_overview(case_id):
    """
    顯示特定案件的詳細成本統計
    
    Args:
        case_id: 案件ID
    """
    # 獲取案件統計資料
    case_stats = get_case_statistics(case_id)
    
    # 顯示案件基本資訊
    # st.subheader(f"📊 {case_stats['CaseName']} (ID: {case_stats['CaseID']})")
    
    # 顯示摘要資訊
    col1, col2, col3 = st.columns(3,border=True)
    
    # 材料成本
    material_cost = case_stats['TotalMaterialCost']
    
    # 人力成本
    labor_cost = case_stats['TotalLaborCost']

    # 計算總成本
    total_cost = material_cost + labor_cost
    
    with col1:
        st.metric("材料成本", f"${material_cost:,}")
    with col2:
        st.metric("人力成本", f"${labor_cost:,}")
    with col3:
        st.metric("總成本", f"${total_cost:,}")
    
    # 顯示材料成本明細
    st.markdown("#### 📦 材料明細")
    if case_stats['Materials']:
        df_materials = pd.DataFrame(case_stats['Materials'])
        st.dataframe(
            df_materials,
            hide_index=True,
            column_config={
                "MaterialID": None,
                "Name": st.column_config.TextColumn("材料名稱"),
                "Unit": st.column_config.TextColumn("單位"),
                "UnitPrice": st.column_config.NumberColumn("單價", format="$%d"),
                "Quantity": st.column_config.NumberColumn("數量"),
                "TotalCost": st.column_config.NumberColumn("總成本", format="$%d"),
            },
            use_container_width=True
        )
    else:
        st.info("此案件尚未使用任何材料")
    
    # 顯示人力工時明細
    st.markdown("#### 👷 工時明細")
    
    if case_stats['Attendances']:
        df_attendances = pd.DataFrame(case_stats['Attendances'])

        # 轉換日期時間格式
        df_attendances['ClockInTime'] = pd.to_datetime(df_attendances['ClockInTime'], format='ISO8601')
        df_attendances['ClockOutTime'] = pd.to_datetime(df_attendances['ClockOutTime'], format='ISO8601')
        
        # 添加日期列以便分組
        df_attendances['Date'] = df_attendances['ClockInTime'].dt.date
        
        # 顯示每日工時統計
        daily_hours = df_attendances.groupby('Date')['WorkHours'].sum().reset_index()
        daily_hours['Date'] = daily_hours['Date'].astype(str)
        
        # 顯示每人工時統計及比例
        user_hours = df_attendances.groupby('UserName')['WorkHours'].sum().reset_index()
        total_hours = user_hours['WorkHours'].sum()
        user_hours['Percentage'] = (user_hours['WorkHours'] / total_hours * 100).round(1)
        user_hours = user_hours.sort_values(by='WorkHours', ascending=False)
        
        # 顯示員工工時統計
        st.dataframe(
            user_hours,
            hide_index=True,
            column_config={
                "UserName": st.column_config.TextColumn("員工名稱"),
                "WorkHours": st.column_config.NumberColumn("工時(小時)"),
                "Percentage": st.column_config.ProgressColumn("工時比例", format="%.1f%%", min_value=0, max_value=100),
            },
            use_container_width=True
        )
        
        # 顯示工時圖表
        # st.markdown("#### 每日工時統計")
        # st.bar_chart(daily_hours.set_index('Date'))
        
        # 顯示原始打卡記錄
        # with st.expander("打卡記錄明細"):
        #     df_attendances['ClockInTime'] = df_attendances['ClockInTime'].dt.strftime("%Y-%m-%d %H:%M")
        #     df_attendances['ClockOutTime'] = df_attendances['ClockOutTime'].dt.strftime("%Y-%m-%d %H:%M")
            
        #     st.dataframe(
        #         df_attendances.drop(columns=['Date']),
        #         hide_index=True,
        #         column_config={
        #             "UserID": st.column_config.TextColumn("員工ID"),
        #             "IsTrained": st.column_config.CheckboxColumn("已訓練"),
        #             "ClockInTime": st.column_config.TextColumn("上班時間"),
        #             "ClockOutTime": st.column_config.TextColumn("下班時間"),
        #             "WorkHours": st.column_config.NumberColumn("工時(小時)"),
        #         },
        #         use_container_width=True
        #     )
    else:
        st.info("此案件尚未有打卡記錄")

    # 顯示施工日誌內容
    st.markdown("#### 📝 施工日誌")
    worklogs=get_worklogs_by_case_id(case_id)
    
    st.dataframe(worklogs,hide_index=True,column_config={
        "UserName":"員工名稱",
        "Content":"施工內容",
        "Progress":"進度",
        "LogTime":"時間"
        })

def display_cost_analysis():
    """
    顯示工程成本分析頁面
    """
    
    try:
        # 先獲取所有案件
        cases = get_cases()
        
        if not cases:
            st.warning("目前沒有案件資料")
            return
        
        # 為每個案件獲取統計資料
        all_cases_stats = []
        
        for case in cases:
            try:
                # 獲取每個案件的統計資料
                case_stats = get_case_statistics(case['CaseID'])
                
                # 建立摘要資料
                summary = {
                    'CaseID': case_stats['CaseID'],
                    'CaseName': case_stats['CaseName'],
                    'Location': case.get('Location', ''),
                    'Status': case.get('Status', 'active'),
                    'TotalMaterialCost': case_stats['TotalMaterialCost'],
                    'TotalLaborCost': case_stats['TotalLaborCost'],
                    'TotalWorkHours': case_stats['TotalWorkHours']
                }
                
                all_cases_stats.append(summary)
                
            except Exception as e:
                # 如果某個案件的統計資料獲取失敗，跳過該案件
                st.warning(f"無法獲取案件 {case['CaseID']} 的統計資料: {e}")
                continue
        
        if not all_cases_stats:
            st.warning("無法獲取任何案件的統計資料")
            return
        
        # 轉換為 DataFrame
        df_cases = pd.DataFrame(all_cases_stats)
        
        # 計算總成本
        df_cases['TotalCost'] = df_cases['TotalMaterialCost'] + df_cases['TotalLaborCost']
        
        # # 側邊欄篩選器
        # st.sidebar.markdown("### 🔍 篩選條件")
        
        # # 工地篩選
        # locations = ['全部'] + sorted(df_cases['Location'].unique().tolist())
        # selected_location = st.sidebar.selectbox("選擇工地", locations)
        
        # # 成本範圍篩選
        # min_cost = int(df_cases['TotalCost'].min())
        # max_cost = int(df_cases['TotalCost'].max())
        # cost_range = st.sidebar.slider(
        #     "成本範圍", 
        #     min_value=min_cost, 
        #     max_value=max_cost, 
        #     value=(min_cost, max_cost),
        #     format="$%d"
        # )
        
        # # 應用篩選
        filtered_df = df_cases.copy()
        # if selected_location != '全部':
        #     filtered_df = filtered_df[filtered_df['Location'] == selected_location]
        
        # filtered_df = filtered_df[
        #     (filtered_df['TotalCost'] >= cost_range[0]) & 
        #     (filtered_df['TotalCost'] <= cost_range[1])
        # ]
        
        # if filtered_df.empty:
        #     st.warning("沒有符合篩選條件的資料")
        #     return
        
        # 總覽儀表板
        # st.markdown("#### 📊 成本總覽")
        col1, col2, col3, col4 = st.columns(4,border=True)
        
        total_projects = len(filtered_df)
        total_cost = filtered_df['TotalCost'].sum()
        avg_cost = filtered_df['TotalCost'].mean()
        highest_cost = filtered_df['TotalCost'].max()
        
        with col1:
            st.metric("工程數量", f"{total_projects} 個")
        with col2:
            st.metric("總成本", f"${total_cost:,.0f}")
        with col3:
            st.metric("平均成本", f"${avg_cost:,.0f}")
        with col4:
            st.metric("最高成本", f"${highest_cost:,.0f}")
        
        st.divider()
        
        # 橫向長條圖 - 工程成本比較
        st.markdown("#### 📊 工程成本比較（材料 vs 人力）")
        
        # 準備圖表資料
        fig = go.Figure()
        
        # 材料成本
        fig.add_trace(go.Bar(
            name='材料成本',
            y=filtered_df['CaseName'],
            x=filtered_df['TotalMaterialCost'],
            orientation='h',
            marker_color='#FF6B6B',
            text=[f'${x:,.0f}' for x in filtered_df['TotalMaterialCost']],
            textposition='inside'
        ))
        
        # 人力成本
        fig.add_trace(go.Bar(
            name='人力成本',
            y=filtered_df['CaseName'],
            x=filtered_df['TotalLaborCost'],
            orientation='h',
            marker_color='#4ECDC4',
            text=[f'${x:,.0f}' for x in filtered_df['TotalLaborCost']],
            textposition='inside'
        ))
        
        fig.update_layout(
            barmode='stack',
            title='各工程成本結構比較',
            xaxis_title='成本金額 ($)',
            yaxis_title='工程名稱',
            height=max(400, len(filtered_df) * 40),  # 根據工程數量調整高度
            showlegend=True,
            yaxis={'categoryorder': 'total ascending'}  # 按總成本排序
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        st.divider()
        
        # 詳細成本控制表格
        st.markdown("#### 📋 詳細成本明細表")
        
        # 準備顯示的資料
        display_df = filtered_df[[
            'CaseName', 'Location', 'Status', 
            'TotalMaterialCost', 'TotalLaborCost', 'TotalCost'
        ]].copy()
        
        # 計算比例
        display_df['MaterialRatio'] = (display_df['TotalMaterialCost'] / display_df['TotalCost'] * 100).round(1)
        display_df['LaborRatio'] = (display_df['TotalLaborCost'] / display_df['TotalCost'] * 100).round(1)
        
        # 按總成本排序
        display_df = display_df.sort_values('TotalCost', ascending=False)
        
        st.dataframe(
            display_df,
            hide_index=True,
            column_config={
                "CaseName": st.column_config.TextColumn("工程名稱", width="medium"),
                "Location": st.column_config.TextColumn("工地位置", width="small"),
                "Status": st.column_config.TextColumn("狀態", width="small"),
                "TotalMaterialCost": st.column_config.NumberColumn(
                    "材料成本", 
                    format="$%d",
                    width="small"
                ),
                "TotalLaborCost": st.column_config.NumberColumn(
                    "人力成本", 
                    format="$%d",
                    width="small"
                ),
                "TotalCost": st.column_config.NumberColumn(
                    "總成本", 
                    format="$%d",
                    width="small"
                ),
                "MaterialRatio": st.column_config.ProgressColumn(
                    "材料比例",
                    format="%.1f%%",
                    min_value=0,
                    max_value=100,
                    width="small"
                ),
                "LaborRatio": st.column_config.ProgressColumn(
                    "人力比例",
                    format="%.1f%%",
                    min_value=0,
                    max_value=100,
                    width="small"
                ),
            },
            use_container_width=True
        )
        
    except Exception as e:
        st.error(f"載入成本資料時發生錯誤: {str(e)}")

tab1, tab2, tab3, tab4 = st.tabs(["⏰ 時效控制", "📊 案件成本","👥 員工總覽", "💰 承攬成本"])

with tab1:
    #證照到期提醒
    st.markdown("#### :bell: 證照到期提醒")

    cert_expired=get_cert_expired()

    if cert_expired:
        
        df_cert_expired=pd.DataFrame(cert_expired)

        df_show=["certificate_name","issue_date","expiry_date","employee_name","days_expired"]

        st.dataframe(df_cert_expired[df_show],hide_index=True,column_config={
            "certificate_name":"證照名稱",
            "issue_date":"發證日",
            "expiry_date":"到期日",
            "employee_name":"員工名稱",
            "days_expired":"超過天數"
        })
    st.divider()
    #機具維護提醒
    st.markdown("#### :bell: 機具維護提醒")
    equipment_maintenance=get_equipment_maintenance()

    if equipment_maintenance:
        df_equipment_maintenance=pd.DataFrame(equipment_maintenance)
        df_show=["EquipmentID","Name","PurchaseDate","NextMaintenance","days_overdue"]
        
        st.dataframe(df_equipment_maintenance[df_show],hide_index=True,column_config={
            "Name":"機具名稱",
            "EquipmentID":"機具ID",
            "PurchaseDate":"購買日期",
            "NextMaintenance":"下次維護日期",
        "days_overdue":"超過天數"
        })
    st.divider()
    # #歸還逾期提醒
    # st.markdown("#### :bell: 借用逾期提醒")
    # st.divider()
    # #出勤異常提醒
    # st.markdown("#### :bell: 出勤異常提醒")


with tab2:
    
    # 獲取所有案件
    cases = get_cases()
    
    # 建立案件選擇器
    if cases:
        case_options = [f"{case['Name']} (ID: {case['CaseID']})" for case in cases]
        selected_case = st.selectbox("選擇案件", case_options)
        
        # 解析選擇的案件ID
        selected_case_id = int(selected_case.split("ID: ")[1].replace(")", ""))
        
        # 顯示選擇的案件統計資料
        display_case_overview(selected_case_id)
    else:
        st.warning("目前沒有可用的案件")
    
    st.divider()

with tab3:
    # 薪資統計部分
    
    # 月份選擇器
    current_year = datetime.datetime.now().year
    months = [
        f"{current_year}-{month:02d}" for month in range(1, 13)
    ]
    selected_month = st.selectbox("選擇月份", months, index=datetime.datetime.now().month-1)
    
    try:
        # 取得選定月份的出勤資料
        attendance = get_attendance_by_month(selected_month)
        
        if attendance and 'employees' in attendance:
            # 分析選定月份的天數
            year, month = map(int, selected_month.split('-'))
            _, last_day = calendar.monthrange(year, month)
            days_in_month = list(range(1, last_day + 1))
            
            # 創建日期欄位
            columns = ["姓名/日期"] + [str(day) for day in days_in_month] + ["合計"]
            
            # 創建表格資料
            table_data = []
            
            # 將每位員工的出勤資料轉換為表格格式
            for emp in attendance['employees']:
                row = [emp['name']]  # 第一欄是員工姓名
                
                # 建立日期對應的工時字典
                hours_by_day = {}
                for record in emp['daily_records']:
                    day = int(record['date'].split('-')[2])  # 取出日期中的「日」
                    hours_by_day[day] = record['hours']
                
                # 填充每一天的工時
                for day in days_in_month:
                    hours = hours_by_day.get(day, 0)
                    # 如果工時大於0，顯示工時值，否則顯示空白
                    row.append(hours if hours > 0 else "")
                
                # 最後一欄是總計
                row.append(str(emp['total_hours']))
                
                table_data.append(row)
            
            # 添加合計行
            total_row = ["合計"]
            for day in days_in_month:
                # 計算每一天的工時總和
                day_total = 0
                for emp in table_data:
                    day_index = day  # 日期對應的索引
                    if day_index < len(emp) and emp[day_index] != "":
                        day_total += float(emp[day_index])
                
                total_row.append(day_total if day_total > 0 else "")
            
            # 最後一格是所有員工總工時
            grand_total = sum(float(emp[-1]) for emp in table_data if emp[-1] != "")
            total_row.append(grand_total)
            
            table_data.append(total_row)
            
            # 創建 DataFrame 並顯示
            df = pd.DataFrame(table_data, columns=columns)
            
            # 顯示表格
            st.markdown(f"### {selected_month} 月份員工出勤表")
            st.dataframe(df, hide_index=True, use_container_width=True)
            
            if st.button("列印薪資單",type="primary"):
                # 建立 Excel 檔案
                excel_file = export_salary_to_excel(attendance, selected_month)
                
                # 提供下載連結
                st.download_button(
                    label="下載薪資單 Excel 檔案",
                    data=excel_file,
                    file_name=f"薪資單_{selected_month}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            # 顯示摘要統計
            # st.markdown("### 出勤摘要統計")
            
            # summary_data = []
            # for emp in attendance['employees']:
            #     # 計算出勤天數 (工時 > 0 的天數)
            #     work_days = sum(1 for day in emp['daily_records'] if day['hours'] > 0)
                
            #     summary_data.append({
            #         "員工姓名": emp['name'],
            #         "總工時": emp['total_hours'],
            #         "出勤天數": work_days,
            #         "換算天數": emp['days_equivalent']
            #     })
            
            # # 顯示摘要資料
            # df_summary = pd.DataFrame(summary_data)
            # st.dataframe(df_summary, hide_index=True, use_container_width=True)
            
        else:
            st.info(f"{selected_month} 月份沒有出勤資料")
    except Exception as e:
        st.error(f"取得資料時發生錯誤: {str(e)}")

with tab4:
    display_cost_analysis()
